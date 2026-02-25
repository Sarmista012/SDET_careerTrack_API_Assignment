from __future__ import annotations
from dataclasses import dataclass
from typing import Any
from openpyxl import load_workbook

@dataclass(frozen=True)
class BookingPayload:
    firstname: str
    lastname: str
    totalprice: int
    depositpaid: bool
    checkin: str
    checkout: str
    additionalneeds: str

def _to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"true", "1", "yes", "y"}

def read_booking_payloads_xlsx(xlsx_path: str, sheet_name: str = "Sheet1", limit: int = 5) -> list[BookingPayload]:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty")

    headers = [str(h).strip().lower() for h in rows[0]]
    idx = {name: headers.index(name) for name in headers}

    required = ["firstname","lastname","totalprice","depositpaid","checkin","checkout","additionalneeds"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    payloads: list[BookingPayload] = []
    for r in rows[1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue

        payloads.append(
            BookingPayload(
                firstname=str(r[idx["firstname"]]).strip(),
                lastname=str(r[idx["lastname"]]).strip(),
                totalprice=int(r[idx["totalprice"]]),
                depositpaid=_to_bool(r[idx["depositpaid"]]),
                checkin=str(r[idx["checkin"]]).strip(),
                checkout=str(r[idx["checkout"]]).strip(),
                additionalneeds=str(r[idx["additionalneeds"]]).strip(),
            )
        )
        if len(payloads) >= limit:
            break

    if len(payloads) < limit:
        raise ValueError(f"Need {limit} payload rows, found {len(payloads)}")

    return payloads
